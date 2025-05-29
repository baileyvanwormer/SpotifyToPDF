from celery_worker import celery_app
import os
from fpdf import FPDF
import spotipy
from spotipy.oauth2 import SpotifyOAuth
from datetime import datetime
import pandas as pd
import openpyxl
import json
import redis
import time
from utils import upload_to_s3


def get_spotify_client(session_token):
    r = redis.Redis.from_url(os.getenv("REDIS_URL"))
    redis_key = f"session:{session_token}"
    token_info_raw = r.get(redis_key)

    if not token_info_raw:
        print("❌ No token_info found in Redis for session_token")
        return None

    token_info = json.loads(token_info_raw)
    access_token = token_info.get("access_token")

    # Refresh token if needed
    if token_info.get("expires_at") and token_info["expires_at"] - int(time.time()) < 60:
        print("🔁 Refreshing expired Spotify token...")
        sp_oauth = SpotifyOAuth(
            client_id=os.getenv("SPOTIPY_CLIENT_ID"),
            client_secret=os.getenv("SPOTIPY_CLIENT_SECRET"),
            redirect_uri=os.getenv("SPOTIPY_REDIRECT_URI")
        )
        token_info = sp_oauth.refresh_access_token(token_info["refresh_token"])
        access_token = token_info["access_token"]
        r.set(redis_key, json.dumps(token_info), ex=3600)

    return spotipy.Spotify(auth=access_token)

@celery_app.task(name="tasks.generate_pdf")
def generate_pdf(session_token, include_liked, playlist_ids, liked_limit):
    sp = get_spotify_client(session_token)
    if not sp:
        return {"status": "failed", "reason": "invalid token"}

    font_dir = os.path.join(os.path.dirname(__file__), "fonts")
    pdf = FPDF()
    pdf.add_page()
    print("📂 Font dir contents:", os.listdir(font_dir))
    print("📍 Font path:", os.path.join(font_dir, "NotoSans-Medium.ttf"))
    pdf.add_font("NotoSans", "", os.path.join(font_dir, "NotoSans-Medium.ttf"), uni=True)
    pdf.add_font("NotoSans", "B", os.path.join(font_dir, "NotoSans-Bold.ttf"), uni=True)
    pdf.add_font("NotoSans", "I", os.path.join(font_dir, "NotoSans-Italic.ttf"), uni=True)

    if include_liked:
        liked_tracks = []
        MAX_SPOTIFY_LIMIT = 50
        fetched = 0

        while fetched < liked_limit:
            to_fetch = min(MAX_SPOTIFY_LIMIT, liked_limit - fetched)
            results = sp.current_user_saved_tracks(limit=to_fetch, offset=fetched)
            for item in results['items']:
                track = item['track']
                artist = track['artists'][0]['name']
                song = track['name']
                liked_tracks.append((artist, song))
            fetched += to_fetch
            if not results.get('next'):
                break

        now = datetime.now().strftime("%B %d, %Y at %I:%M %p")
        pdf.set_font("NotoSans", "I", 10)
        pdf.cell(0, 10, f"Exported on {now}", ln=True)
        pdf.ln(5)

        pdf.set_font("NotoSans", "B", 16)
        pdf.cell(0, 10, "🎵 Liked Songs", ln=True)
        pdf.ln(5)

        for i, (artist, song) in enumerate(liked_tracks):
            pdf.set_font("NotoSans", "", 12)
            pdf.multi_cell(0, 10, f"{i+1}. {artist} – {song}")
            pdf.ln(1)

    for pid in playlist_ids:
        playlist = sp.playlist(pid)
        tracks_data = sp.playlist_tracks(pid)
        tracks = []
        for item in tracks_data['items']:
            track = item['track']
            if track:
                artist = track['artists'][0]['name']
                song = track['name']
                tracks.append((artist, song))

        pdf.add_page()
        pdf.set_font("NotoSans", "B", 16)
        pdf.cell(0, 10, f"📂 Playlist: {playlist['name']}", ln=True)
        pdf.ln(5)

        for i, (artist, song) in enumerate(tracks):
            pdf.set_font("NotoSans", "", 12)
            pdf.multi_cell(0, 10, f"{i+1}. {artist} – {song}")
            pdf.ln(1)

    pdf_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "spotify_export.pdf"))
    pdf.output(pdf_path)
    url = upload_to_s3(pdf_path)
    print(f"✅ Returning S3 URL: {url}")
    return url


@celery_app.task(name="tasks.generate_excel")
def generate_excel(session_token, include_liked, playlist_ids, liked_limit):
    print("🔧 Starting Excel task...")
    print("🔑 Session Token (partial):", session_token[:10])

    sp = get_spotify_client(session_token)
    print("sp: ", session_token)
    if not sp:
        return {"status": "failed", "reason": "invalid token"}

    try:
        rows = []
        MAX_SPOTIFY_LIMIT = 50
        fetched = 0

        if include_liked:
            print("🎶 Fetching liked songs...")
            while fetched < liked_limit:
                to_fetch = min(MAX_SPOTIFY_LIMIT, liked_limit - fetched)
                results = sp.current_user_saved_tracks(limit=to_fetch, offset=fetched)
                print(f"📦 Fetched {len(results['items'])} tracks at offset {fetched}")
                for item in results['items']:
                    track = item['track']
                    rows.append({
                        "Source": "Liked Songs",
                        "Artist": track['artists'][0]['name'],
                        "Track": track['name'],
                        "Album": track['album']['name'],
                        "Duration (min)": "{:d}:{:02d}".format(
                            (track['duration_ms'] // 60000),
                            (track['duration_ms'] // 1000) % 60
                        )
                    })
                fetched += to_fetch
                if not results.get('next'):
                    break

        for pid in playlist_ids:
            print(f"📁 Fetching playlist: {pid}")
            playlist = sp.playlist(pid)
            playlist_name = playlist['name']
            tracks_data = sp.playlist_tracks(pid)
            for item in tracks_data['items']:
                track = item['track']
                if track:
                    rows.append({
                        "Source": playlist_name,
                        "Artist": track['artists'][0]['name'],
                        "Track": track['name'],
                        "Album": track['album']['name'],
                        "Duration (min)": "{:d}:{:02d}".format(
                            (track['duration_ms'] // 60000),
                            (track['duration_ms'] // 1000) % 60
                        )
                    })

        if not rows:
            print("⚠️ No rows collected — skipping Excel export.")
            return {"status": "failed", "reason": "no data"}

        import uuid
        filename = f"spotify_export_{uuid.uuid4().hex}.xlsx"
        xlsx_path = os.path.join(os.path.dirname(__file__), filename)
        print(f"💾 Writing Excel to: {xlsx_path}")

        df = pd.DataFrame(rows, columns=["Source", "Artist", "Track", "Album", "Duration (min)"])
        df.to_excel(xlsx_path, index=False, engine='openpyxl')

        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        column_widths = {"A": 20, "B": 25, "C": 35, "D": 30, "E": 15}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        wb.save(xlsx_path)
        wb.close()

        print(f"✅ Excel complete: {xlsx_path} ({os.path.getsize(xlsx_path)} bytes)")
        url = upload_to_s3(xlsx_path)
        print(f"✅ Returning S3 URL: {url}")
        return url

    except Exception as e:
        print("❌ Excel generation failed:", e)
        return {"status": "failed", "reason": str(e)}